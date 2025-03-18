import openpyxl
import os

# Define the path to the Excel file
excel_file = "C:\\Users\\ManjulaAlagarsamy\\PycharmProjects\\PythonProject\\sheetutilities\\Truewave.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(excel_file)

def selectsheet(sheetname):
    """Selects the specified sheet from the Excel file."""
    global worksheet
    worksheet = workbook[sheetname]

def rowcount():
    """Returns the number of non-empty rows in the selected sheet."""
    return worksheet.max_row

def readdata(cell):
    """Reads data from a specific cell."""
    return worksheet[cell].value

def writedata(row, column, data):
    """Writes data to a specific cell and saves the file."""
    worksheet.cell(row=row, column=column, value=data)
    workbook.save(excel_file)  # Save changes to the Excel file

# Example usage:
selectsheet("Sheet1")  # Select the sheet

print(readdata("A1"))  # Read data from A1
print(f"Total Rows: {rowcount()}")  # Get row count

writedata(2, 2, "Hello Excel")  # Write data to B2
